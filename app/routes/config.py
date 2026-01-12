from fastapi import APIRouter, UploadFile, File, HTTPException
from typing import Optional, List
from pydantic import BaseModel
import os
from pathlib import Path

from ..models import (
    Function, Team, TeamWithFunction, Tool, Capability, ConfigData,
    FunctionCreate, TeamCreate, ToolCreate, CapabilityCreate
)
from .. import crud

router = APIRouter(prefix="/api/config", tags=["config"])

UPLOADS_DIR = Path(__file__).parent.parent.parent / "uploads"


# Update models for PUT requests
class ItemUpdate(BaseModel):
    name: str


class TeamUpdate(BaseModel):
    name: str
    function_id: int


class MoveEntriesRequest(BaseModel):
    target_team_id: Optional[int] = None  # None means move to function level (no team)


@router.get("", response_model=ConfigData)
async def get_all_config():
    """Get all configuration data (functions, teams, tools, capabilities)."""
    return ConfigData(
        functions=crud.get_functions(),
        teams=crud.get_teams(),
        tools=crud.get_tools(),
        capabilities=crud.get_capabilities()
    )


# ============ Functions CRUD ============

@router.get("/functions", response_model=List[dict])
async def get_functions():
    """List all functions."""
    return crud.get_functions()


@router.post("/functions")
async def create_function(func: FunctionCreate):
    """Create a new function."""
    return crud.create_function(func)


@router.put("/functions/{function_id}")
async def update_function(function_id: int, data: ItemUpdate):
    """Update a function."""
    result = crud.update_function(function_id, data.name)
    if not result:
        raise HTTPException(status_code=404, detail="Function not found")
    return result


@router.delete("/functions/{function_id}")
async def delete_function(function_id: int):
    """Delete a function (only if no teams or responses reference it)."""
    deleted = crud.delete_function(function_id)
    if not deleted:
        raise HTTPException(status_code=400, detail="Cannot delete: function has teams or responses")
    return {"message": "Function deleted"}


# ============ Teams CRUD ============

@router.get("/teams", response_model=List[dict])
async def get_teams(function_id: Optional[int] = None):
    """List all teams, optionally filtered by function_id."""
    return crud.get_teams(function_id=function_id)


@router.post("/teams")
async def create_team(team: TeamCreate):
    """Create a new team."""
    return crud.create_team(team)


@router.put("/teams/{team_id}")
async def update_team(team_id: int, data: TeamUpdate):
    """Update a team."""
    result = crud.update_team(team_id, data.name, data.function_id)
    if not result:
        raise HTTPException(status_code=404, detail="Team not found")
    return result


@router.delete("/teams/{team_id}")
async def delete_team(team_id: int):
    """Delete a team (only if no responses reference it)."""
    deleted = crud.delete_team(team_id)
    if not deleted:
        raise HTTPException(status_code=400, detail="Cannot delete: team has responses")
    return {"message": "Team deleted"}


@router.get("/teams/{team_id}/entries")
async def get_team_entries_info(team_id: int):
    """Get team entry count and sibling teams for move operation."""
    result = crud.get_team_entry_count(team_id)
    if not result:
        raise HTTPException(status_code=404, detail="Team not found")
    return result


@router.post("/teams/{team_id}/move-and-delete")
async def move_entries_and_delete_team(team_id: int, request: MoveEntriesRequest):
    """Move all entries to another team (or function level) and delete the team."""
    result = crud.move_team_entries_and_delete(team_id, request.target_team_id)
    if not result.get('success'):
        raise HTTPException(status_code=400, detail=result.get('error', 'Operation failed'))
    return result


# ============ Tools CRUD ============

@router.get("/tools", response_model=List[dict])
async def get_tools():
    """List all tools."""
    return crud.get_tools()


@router.post("/tools")
async def create_tool(tool: ToolCreate):
    """Create a new tool."""
    return crud.create_tool(tool)


@router.put("/tools/{tool_id}")
async def update_tool(tool_id: int, data: ItemUpdate):
    """Update a tool."""
    result = crud.update_tool(tool_id, data.name)
    if not result:
        raise HTTPException(status_code=404, detail="Tool not found")
    return result


@router.delete("/tools/{tool_id}")
async def delete_tool(tool_id: int):
    """Delete a tool."""
    deleted = crud.delete_tool(tool_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Tool not found")
    return {"message": "Tool deleted"}


# ============ Capabilities CRUD ============

@router.get("/capabilities", response_model=List[dict])
async def get_capabilities():
    """List all capabilities."""
    return crud.get_capabilities()


@router.post("/capabilities")
async def create_capability(cap: CapabilityCreate):
    """Create a new capability."""
    return crud.create_capability(cap)


@router.put("/capabilities/{capability_id}")
async def update_capability(capability_id: int, data: ItemUpdate):
    """Update a capability."""
    result = crud.update_capability(capability_id, data.name)
    if not result:
        raise HTTPException(status_code=404, detail="Capability not found")
    return result


@router.delete("/capabilities/{capability_id}")
async def delete_capability(capability_id: int):
    """Delete a capability (only if no responses reference it)."""
    deleted = crud.delete_capability(capability_id)
    if not deleted:
        raise HTTPException(status_code=400, detail="Cannot delete: capability has responses")
    return {"message": "Capability deleted"}


@router.post("/upload")
async def upload_config(file: UploadFile = File(...)):
    """Upload Excel file to update configuration."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    try:
        import openpyxl

        # Save the uploaded file
        os.makedirs(UPLOADS_DIR, exist_ok=True)
        file_path = UPLOADS_DIR / "config.xlsx"

        contents = await file.read()
        with open(file_path, "wb") as f:
            f.write(contents)

        # Parse the Excel file
        wb = openpyxl.load_workbook(file_path)

        functions = []
        teams = {}
        tools = []
        capabilities = []

        # Parse Functions sheet
        if "Functions" in wb.sheetnames:
            ws = wb["Functions"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    functions.append(str(row[0]).strip())

        # Parse Teams sheet
        if "Teams" in wb.sheetnames:
            ws = wb["Teams"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    func_name = str(row[0]).strip()
                    team_name = str(row[1]).strip()
                    if func_name not in teams:
                        teams[func_name] = []
                    teams[func_name].append(team_name)

        # Parse Tools sheet
        if "Tools" in wb.sheetnames:
            ws = wb["Tools"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    tools.append(str(row[0]).strip())

        # Parse Capabilities sheet
        if "Capabilities" in wb.sheetnames:
            ws = wb["Capabilities"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    name = str(row[0]).strip()
                    icon = str(row[1]).strip() if len(row) > 1 and row[1] else None
                    capabilities.append((name, icon))

        # Update the database
        crud.clear_and_reload_config(functions, teams, tools, capabilities)

        return {
            "message": "Configuration uploaded successfully",
            "functions_count": len(functions),
            "teams_count": sum(len(t) for t in teams.values()),
            "tools_count": len(tools),
            "capabilities_count": len(capabilities)
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
